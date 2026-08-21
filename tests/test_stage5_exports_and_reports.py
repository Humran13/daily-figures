"""
Stage 5, Section 10: report/PDF quantity language. Dispatch's export (CSV/
XLSX/PDF) no longer leads with raw Cartons/Packs/Pieces/Total-Pieces
columns — a single business-friendly "Quantity" column per product,
reusing the same formatter as Returns/Production/Daily Figures (see
webapp/services/quantity_format.py). This is a presentation correction,
not a conversion-logic rewrite: exact values are unchanged, only how they're
displayed. CSV must stay pure, machine-readable data throughout.
"""
import pytest


@pytest.fixture
def super_admin(login_as):
    return login_as("root", "password123", "super_admin")


@pytest.fixture
def setup(client, super_admin):
    product = client.post("/api/admin/products", json={"name": "Compact Corporate Test"}).get_json()
    client.post(f"/api/admin/products/{product['id']}/packaging-rules", json={
        "cartons_to_packs": 10, "packs_to_pieces": 10,
    })
    category = client.post("/api/admin/sales-categories", json={"name": "Test Category"}).get_json()
    customer = client.post("/api/admin/customers", json={"name": "Dalca", "sales_category_id": category["id"]}).get_json()
    return {"product": product, "customer": customer, "category": category}


def _finalize_dispatch(client, product_id, customer_id, date, shift, cartons, packs, pieces, number):
    created = client.post("/api/dispatches", json={
        "dispatch_number": number, "date": date, "shift": shift, "customer_id": customer_id,
        "lines": [{"product_id": product_id, "cartons": cartons, "packs": packs, "pieces": pieces}],
    }).get_json()
    client.post(f"/api/dispatches/{created['id']}/finalize")
    return created


# ---------- Dispatch export: presentation correction ----------

def test_dispatch_csv_no_longer_has_separate_cartons_packs_pieces_columns(client, setup):
    _finalize_dispatch(client, setup["product"]["id"], setup["customer"]["id"], "2026-07-28", "Day", 2, 3, 4, "RPT-1")
    res = client.get("/api/dispatches/export.csv")
    text = res.data.decode()
    header_line = text.splitlines()[4]  # title/generated/filters/blank/header
    assert "Cartons" not in header_line
    assert "Packs" not in header_line
    assert "Pieces" not in header_line
    assert "Total Pieces" not in header_line
    assert "Quantity" in header_line


def test_dispatch_csv_shows_business_friendly_quantity_string(client, setup):
    _finalize_dispatch(client, setup["product"]["id"], setup["customer"]["id"], "2026-07-28", "Day", 2, 3, 4, "RPT-2")
    res = client.get("/api/dispatches/export.csv")
    assert b"2.34 Ctns" in res.data


def test_dispatch_no_pack_tier_product_shows_book_style_point_notation(client, setup):
    """Final pre-deployment correction: a no-pack-tier product's book-style
    quantity now uses the same "C.PP Ctns" point notation as pack-tier
    products — the loose-piece remainder positioned after the point, not
    the older "Xc Ypc" form."""
    kingmax = client.post("/api/admin/products", json={"name": "KingMax Test"}).get_json()
    client.post(f"/api/admin/products/{kingmax['id']}/packaging-rules", json={"carton_to_pieces": 60})
    _finalize_dispatch(client, kingmax["id"], setup["customer"]["id"], "2026-07-28", "Day", 2, 0, 5, "RPT-3")
    res = client.get("/api/dispatches/export.csv")
    assert b"2.05 Ctns" in res.data


def test_dispatch_csv_has_no_total_pieces_or_base_unit_row_at_all(client, setup):
    """Stage 5 correction: normal, business-facing exports must never show a
    raw pieces/base-unit total, not even as a labeled diagnostic row — there
    is no separate admin-only diagnostic export in this app to put one in,
    so the correct fix is simply not emitting it here."""
    _finalize_dispatch(client, setup["product"]["id"], setup["customer"]["id"], "2026-07-28", "Day", 1, 0, 0, "RPT-4")
    res = client.get("/api/dispatches/export.csv")
    text = res.data.decode()
    assert "TOTAL" not in text
    assert "diagnostic" not in text.lower()
    # exactly: title, generated-by, filters, blank, header, one data row —
    # nothing trailing.
    lines = [ln for ln in text.splitlines() if ln != ""]
    assert len(lines) == 5


def test_dispatch_xlsx_has_no_trailing_total_row(client, setup):
    _finalize_dispatch(client, setup["product"]["id"], setup["customer"]["id"], "2026-07-28", "Day", 1, 0, 0, "RPT-4B")
    res = client.get("/api/dispatches/export.xlsx")
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(res.data))
    ws = wb.active
    values = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
    non_empty_rows = [row for row in values if any(v not in (None, "") for v in row)]
    # title / generated-by / filters / header / one data row — nothing after.
    assert len(non_empty_rows) == 5
    last_row = non_empty_rows[-1]
    assert not any(str(v).upper().startswith("TOTAL") for v in last_row if v)


def test_dispatch_export_route_never_passes_totals_to_build_export():
    """Source-level guard: the export route must not construct a totals
    dict at all — the removal must be structural, not just empty data,
    since build_pdf/build_xlsx/build_csv would render any totals dict
    they're given, including in PDF/print output."""
    from pathlib import Path
    source = Path(__file__).resolve().parent.parent.joinpath("webapp", "routes", "dispatches.py").read_text(encoding="utf-8")
    fn_start = source.index("def export_dispatches(fmt):")
    fn_end = source.index("\n\n\n", fn_start)
    body = source[fn_start:fn_end]
    assert "totals=" not in body
    assert "total_pieces" not in body


def test_dispatch_pdf_and_xlsx_still_work_after_column_rework(client, setup):
    _finalize_dispatch(client, setup["product"]["id"], setup["customer"]["id"], "2026-07-28", "Day", 1, 0, 0, "RPT-5")
    pdf = client.get("/api/dispatches/export.pdf")
    assert pdf.status_code == 200 and pdf.data.startswith(b"%PDF")
    xlsx = client.get("/api/dispatches/export.xlsx")
    assert xlsx.status_code == 200


# ---------- Daily Figures export: Opening Stock / Closing Stock terminology ----------

def test_daily_figures_csv_uses_opening_stock_and_closing_stock_headers(client, setup):
    client.post("/api/daily-figures", json={
        "product_id": setup["product"]["id"], "date": "2026-07-28", "shift": "Day",
        "opening": {"cartons": 1, "packs": 0, "pieces": 0},
    })
    res = client.get("/api/daily-figures/export.csv")
    text = res.data.decode()
    header_line = text.splitlines()[4]
    assert "Opening Stock" in header_line
    assert "Closing Stock" in header_line
    assert "Opening Cartons" not in header_line
    assert "Closing Cartons" not in header_line


def test_daily_figures_csv_quantity_strings_are_business_friendly(client, setup):
    client.post("/api/daily-figures", json={
        "product_id": setup["product"]["id"], "date": "2026-07-28", "shift": "Day",
        "opening": {"cartons": 1, "packs": 2, "pieces": 3},
    })
    res = client.get("/api/daily-figures/export.csv")
    assert b"1.23 Ctns" in res.data


# ---------- CSV stays machine-readable (no decorative branding rows) ----------

def test_dispatch_csv_has_no_decorative_branding_row(client, setup):
    client.patch("/api/admin/company-settings", json={"display_name": "Acme Tissue Co."})
    _finalize_dispatch(client, setup["product"]["id"], setup["customer"]["id"], "2026-07-28", "Day", 1, 0, 0, "RPT-6")
    res = client.get("/api/dispatches/export.csv")
    text = res.data.decode()
    # Row 1 is always the report title (existing convention) — branding
    # must never displace it or insert an extra row ahead of it.
    assert text.splitlines()[0] == "Dispatch Transactions"


def test_returns_and_production_csv_also_have_no_decorative_branding_row(client, setup):
    client.patch("/api/admin/company-settings", json={"display_name": "Acme Tissue Co."})
    created = client.post("/api/returns", json={
        "date": "2026-07-28", "customer_id": setup["customer"]["id"],
        "lines": [{"product_id": setup["product"]["id"], "cartons": 1, "packs": 0, "pieces": 0}],
    }).get_json()
    client.post(f"/api/returns/{created['id']}/finalize")
    res = client.get("/api/returns/export.csv")
    assert res.data.decode().splitlines()[0] == "Returns"


# ---------- pagination on the new list endpoints ----------

def test_returns_list_respects_limit_and_offset(client, setup):
    for i in range(3):
        client.post("/api/returns", json={
            "date": "2026-07-28",
            "lines": [{"product_id": setup["product"]["id"], "cartons": 1, "packs": 0, "pieces": 0}],
        })
    page1 = client.get("/api/returns?limit=2&offset=0").get_json()
    page2 = client.get("/api/returns?limit=2&offset=2").get_json()
    assert page1["total"] == 3
    assert len(page1["results"]) == 2
    assert len(page2["results"]) == 1
    assert {r["id"] for r in page1["results"]}.isdisjoint({r["id"] for r in page2["results"]})


def test_production_list_respects_limit_and_offset(client, setup):
    for i in range(3):
        client.post("/api/production", json={
            "date": "2026-07-28", "shift": "Day",
            "lines": [{"product_id": setup["product"]["id"], "cartons": 1, "packs": 0, "pieces": 0}],
        })
    page1 = client.get("/api/production?limit=2&offset=0").get_json()
    assert page1["total"] == 3
    assert len(page1["results"]) == 2
