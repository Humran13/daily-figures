"""
Final Stage 2 correction: the product_id filter in
_filtered_dispatch_query previously did `query.join(DispatchLine)`, which
multiplies a Dispatch row once per matching DispatchLine — corrupting
count()/pagination/ordering in GET /api/dispatches and repeating every
export format's rows for any dispatch with more than one line matching the
filtered product. Nothing in the schema prevents a dispatch from having
more than one line for the same product (see webapp/models/dispatch.py —
DispatchLine has no unique constraint on (dispatch_id, product_id); only
the New Dispatch UI nudges against it client-side), so this was a real,
reachable bug, not a hypothetical one.

Fixed with a correlated EXISTS subquery instead of a join — see
_filtered_dispatch_query in webapp/routes/dispatches.py.
"""
import pytest


@pytest.fixture
def setup(client, login_as):
    login_as("root", "password123", "super_admin")
    product_a = client.post("/api/admin/products", json={"name": "Dup Product A"}).get_json()
    client.post(f"/api/admin/products/{product_a['id']}/packaging-rules", json={
        "cartons_to_packs": 10, "packs_to_pieces": 10,
    })
    product_b = client.post("/api/admin/products", json={"name": "Dup Product B"}).get_json()
    client.post(f"/api/admin/products/{product_b['id']}/packaging-rules", json={
        "cartons_to_packs": 10, "packs_to_pieces": 10,
    })
    category = client.post("/api/admin/sales-categories", json={"name": "Dup Category"}).get_json()
    customer = client.post("/api/admin/customers", json={
        "name": "Dup Customer", "sales_category_id": category["id"],
    }).get_json()
    return {"product_a": product_a, "product_b": product_b, "category": category, "customer": customer}


def _dispatch_with_lines(client, customer_id, number, lines, date="2026-07-28"):
    return client.post("/api/dispatches", json={
        "dispatch_number": number, "date": date, "shift": "Day", "customer_id": customer_id,
        "lines": lines,
    }).get_json()


# ---------- one dispatch, multiple lines for the SAME product ----------

def test_dispatch_with_duplicate_product_lines_appears_once_in_list(client, setup):
    pid = setup["product_a"]["id"]
    d = _dispatch_with_lines(client, setup["customer"]["id"], "DUP-SAME", lines=[
        {"product_id": pid, "cartons": 1, "packs": 0, "pieces": 0},
        {"product_id": pid, "cartons": 2, "packs": 0, "pieces": 0},
    ])
    assert len(d["lines"]) == 2  # backend has no unique (dispatch_id, product_id) constraint — see module docstring

    res = client.get(f"/api/dispatches?product_id={pid}")
    data = res.get_json()
    matches = [r for r in data["results"] if r["id"] == d["id"]]
    assert len(matches) == 1
    assert data["total"] == 1


def test_dispatch_with_duplicate_product_lines_not_repeated_in_csv_export(client, setup):
    pid = setup["product_a"]["id"]
    d = _dispatch_with_lines(client, setup["customer"]["id"], "DUP-CSV", lines=[
        {"product_id": pid, "cartons": 1, "packs": 0, "pieces": 0},
        {"product_id": pid, "cartons": 2, "packs": 0, "pieces": 0},
    ])
    res = client.get(f"/api/dispatches/export.csv?product_id={pid}")
    assert res.status_code == 200
    body = res.get_data(as_text=True)
    # Exactly 2 rows for this dispatch (its 2 real lines) — never 4, which
    # is what the join(DispatchLine) bug would have produced.
    assert body.count("DUP-CSV") == 2


def test_dispatch_with_duplicate_product_lines_not_repeated_in_excel_export(client, setup):
    pid = setup["product_a"]["id"]
    d = _dispatch_with_lines(client, setup["customer"]["id"], "DUP-XLSX", lines=[
        {"product_id": pid, "cartons": 1, "packs": 0, "pieces": 0},
        {"product_id": pid, "cartons": 2, "packs": 0, "pieces": 0},
    ])
    res = client.get(f"/api/dispatches/export.xlsx?product_id={pid}")
    assert res.status_code == 200

    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(res.data))
    ws = wb.active
    matches = sum(1 for row in ws.iter_rows(values_only=True) if row and row[2] == "DUP-XLSX")
    assert matches == 2


def test_dispatch_with_duplicate_product_lines_pdf_export_succeeds(client, setup):
    """
    The PDF builder consumes the exact same `rows` list constructed once in
    export_dispatches() as CSV/Excel — proven duplicate-free above — so this
    is a functional smoke check that the same code path doesn't error for
    the PDF format, not a second independent row-count proof.
    """
    pid = setup["product_a"]["id"]
    _dispatch_with_lines(client, setup["customer"]["id"], "DUP-PDF", lines=[
        {"product_id": pid, "cartons": 1, "packs": 0, "pieces": 0},
        {"product_id": pid, "cartons": 2, "packs": 0, "pieces": 0},
    ])
    res = client.get(f"/api/dispatches/export.pdf?product_id={pid}")
    assert res.status_code == 200
    assert res.mimetype == "application/pdf"


# ---------- one dispatch, several DIFFERENT products ----------

def test_dispatch_with_multiple_different_products_matches_each_filter_once(client, setup):
    pid_a, pid_b = setup["product_a"]["id"], setup["product_b"]["id"]
    d = _dispatch_with_lines(client, setup["customer"]["id"], "MULTI-PROD", lines=[
        {"product_id": pid_a, "cartons": 1, "packs": 0, "pieces": 0},
        {"product_id": pid_b, "cartons": 1, "packs": 0, "pieces": 0},
    ])

    for pid in (pid_a, pid_b):
        res = client.get(f"/api/dispatches?product_id={pid}")
        data = res.get_json()
        assert data["total"] == 1
        assert len([r for r in data["results"] if r["id"] == d["id"]]) == 1


# ---------- combined with other filters ----------

def test_product_filter_combined_with_customer_name_filter(client, setup):
    pid = setup["product_a"]["id"]
    d = _dispatch_with_lines(client, setup["customer"]["id"], "COMBO-PROD-CUST", lines=[
        {"product_id": pid, "cartons": 1, "packs": 0, "pieces": 0},
        {"product_id": pid, "cartons": 1, "packs": 0, "pieces": 0},
    ])
    res = client.get(f"/api/dispatches?product_id={pid}&customer_name=Dup Customer")
    data = res.get_json()
    assert data["total"] == 1
    assert data["results"][0]["id"] == d["id"]


def test_product_filter_combined_with_status_and_category(client, setup):
    pid = setup["product_a"]["id"]
    d = _dispatch_with_lines(client, setup["customer"]["id"], "COMBO-STATUS", lines=[
        {"product_id": pid, "cartons": 1, "packs": 0, "pieces": 0},
        {"product_id": pid, "cartons": 1, "packs": 0, "pieces": 0},
    ])
    client.post(f"/api/dispatches/{d['id']}/finalize")

    res = client.get(
        f"/api/dispatches?product_id={pid}&status=finalized&sales_category_id={setup['category']['id']}"
    )
    data = res.get_json()
    assert data["total"] == 1
    assert data["results"][0]["id"] == d["id"]


# ---------- pagination stays correct with the product filter ----------

def test_pagination_total_correct_with_product_filter(client, setup):
    pid = setup["product_a"]["id"]
    # 3 distinct dispatches, one of which has duplicate lines for the same
    # product — total must still be 3, not inflated.
    _dispatch_with_lines(client, setup["customer"]["id"], "PGN-A", lines=[
        {"product_id": pid, "cartons": 1, "packs": 0, "pieces": 0},
        {"product_id": pid, "cartons": 1, "packs": 0, "pieces": 0},
    ])
    _dispatch_with_lines(client, setup["customer"]["id"], "PGN-B", lines=[
        {"product_id": pid, "cartons": 1, "packs": 0, "pieces": 0},
    ])
    _dispatch_with_lines(client, setup["customer"]["id"], "PGN-C", lines=[
        {"product_id": pid, "cartons": 1, "packs": 0, "pieces": 0},
    ])

    res = client.get(f"/api/dispatches?product_id={pid}&limit=2&offset=0")
    data = res.get_json()
    assert data["total"] == 3
    assert len(data["results"]) == 2

    res2 = client.get(f"/api/dispatches?product_id={pid}&limit=2&offset=2")
    assert len(res2.get_json()["results"]) == 1


def test_stable_ordering_across_pages_with_product_filter(client, setup):
    pid = setup["product_a"]["id"]
    numbers = ["ORD-1", "ORD-2", "ORD-3", "ORD-4"]
    for n in numbers:
        _dispatch_with_lines(client, setup["customer"]["id"], n, lines=[
            {"product_id": pid, "cartons": 1, "packs": 0, "pieces": 0},
            {"product_id": pid, "cartons": 1, "packs": 0, "pieces": 0},
        ])

    full = client.get(f"/api/dispatches?product_id={pid}&limit=200").get_json()["results"]
    full_ids = [r["id"] for r in full]

    page1 = client.get(f"/api/dispatches?product_id={pid}&limit=2&offset=0").get_json()["results"]
    page2 = client.get(f"/api/dispatches?product_id={pid}&limit=2&offset=2").get_json()["results"]
    paged_ids = [r["id"] for r in page1] + [r["id"] for r in page2]

    assert paged_ids == full_ids[:4]
    assert len(set(paged_ids)) == len(paged_ids)  # no dispatch repeated across pages
