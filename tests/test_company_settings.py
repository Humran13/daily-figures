"""
Stage 3: Company Settings / white-label branding — persistence,
super_admin-only write access, logo upload security, audit logging, and
export/branding-display integration.
"""
import io
import os

import pytest
from PIL import Image


def _png_bytes(size=(20, 20), color="red"):
    buf = io.BytesIO()
    Image.new("RGB", size, color=color).save(buf, format="PNG")
    buf.seek(0)
    return buf


def _jpeg_bytes():
    buf = io.BytesIO()
    Image.new("RGB", (20, 20), color="blue").save(buf, format="JPEG")
    buf.seek(0)
    return buf


def _webp_bytes():
    buf = io.BytesIO()
    Image.new("RGB", (20, 20), color="green").save(buf, format="WEBP")
    buf.seek(0)
    return buf


# ---------- permissions ----------

def test_public_branding_accessible_unauthenticated(client):
    res = client.get("/api/branding")
    assert res.status_code == 200
    data = res.get_json()
    assert data["display_name"] == "Daily Figures"
    assert data["logo_url"] is None


def test_full_settings_requires_login(client):
    res = client.get("/api/admin/company-settings")
    assert res.status_code == 401


@pytest.mark.parametrize("role", ["manager", "operator", "viewer"])
def test_full_settings_read_requires_super_admin(client, login_as, role):
    login_as(f"reader_{role}", "password123", role)
    res = client.get("/api/admin/company-settings")
    assert res.status_code == 403


@pytest.mark.parametrize("role", ["manager", "operator", "viewer"])
def test_update_settings_requires_super_admin(client, login_as, role):
    login_as(f"writer_{role}", "password123", role)
    res = client.patch("/api/admin/company-settings", json={"display_name": "Hacked"})
    assert res.status_code == 403


@pytest.mark.parametrize("role", ["manager", "operator", "viewer"])
def test_logo_upload_requires_super_admin(client, login_as, role):
    login_as(f"uploader_{role}", "password123", role)
    res = client.post("/api/admin/company-settings/logo",
                       data={"logo": (_png_bytes(), "logo.png")}, content_type="multipart/form-data")
    assert res.status_code == 403


@pytest.mark.parametrize("role", ["manager", "operator", "viewer"])
def test_logo_remove_requires_super_admin(client, login_as, role):
    login_as(f"remover_{role}", "password123", role)
    res = client.delete("/api/admin/company-settings/logo")
    assert res.status_code == 403


def test_super_admin_can_read_and_update(client, login_as):
    login_as("root", "password123", "super_admin")
    assert client.get("/api/admin/company-settings").status_code == 200
    res = client.patch("/api/admin/company-settings", json={"display_name": "Acme Foods"})
    assert res.status_code == 200
    assert res.get_json()["display_name"] == "Acme Foods"


# ---------- persistence ----------

def test_settings_persist_across_requests(client, login_as):
    login_as("root", "password123", "super_admin")
    client.patch("/api/admin/company-settings", json={
        "display_name": "Persisted Co", "phone": "+256700000000", "currency_code": "UGX",
    })
    res = client.get("/api/admin/company-settings")
    data = res.get_json()
    assert data["display_name"] == "Persisted Co"
    assert data["phone"] == "+256700000000"
    assert data["currency_code"] == "UGX"


def test_default_settings_seeded_safely(client, login_as):
    login_as("root", "password123", "super_admin")
    data = client.get("/api/admin/company-settings").get_json()
    assert data["display_name"] == "Daily Figures"
    for field in ("legal_name", "address", "phone", "email", "website",
                  "currency_code", "tax_registration_number", "report_footer_text",
                  "primary_contact_name", "logo_url"):
        assert data[field] is None


# ---------- validation ----------

def test_display_name_cannot_be_blanked(client, login_as):
    login_as("root", "password123", "super_admin")
    res = client.patch("/api/admin/company-settings", json={"display_name": "   "})
    assert res.status_code == 400


def test_invalid_email_rejected(client, login_as):
    login_as("root", "password123", "super_admin")
    res = client.patch("/api/admin/company-settings", json={"email": "not-an-email"})
    assert res.status_code == 400


def test_invalid_website_rejected(client, login_as):
    login_as("root", "password123", "super_admin")
    res = client.patch("/api/admin/company-settings", json={"website": "ftp://not-http.example"})
    assert res.status_code == 400


def test_valid_email_and_website_accepted(client, login_as):
    login_as("root", "password123", "super_admin")
    res = client.patch("/api/admin/company-settings", json={
        "email": "info@acme.example", "website": "https://acme.example",
    })
    assert res.status_code == 200


def test_field_length_limit_enforced(client, login_as):
    login_as("root", "password123", "super_admin")
    res = client.patch("/api/admin/company-settings", json={"phone": "1" * 41})
    assert res.status_code == 400


def test_unknown_field_rejected(client, login_as):
    login_as("root", "password123", "super_admin")
    res = client.patch("/api/admin/company-settings", json={"not_a_real_field": "x"})
    assert res.status_code == 400


def test_optional_fields_can_be_cleared(client, login_as):
    login_as("root", "password123", "super_admin")
    client.patch("/api/admin/company-settings", json={"legal_name": "Acme Foods Ltd"})
    res = client.patch("/api/admin/company-settings", json={"legal_name": ""})
    assert res.status_code == 200
    assert res.get_json()["legal_name"] is None


# ---------- logo security ----------

def test_valid_png_upload_succeeds(client, login_as):
    login_as("root", "password123", "super_admin")
    res = client.post("/api/admin/company-settings/logo",
                       data={"logo": (_png_bytes(), "logo.png")}, content_type="multipart/form-data")
    assert res.status_code == 200
    assert res.get_json()["logo_url"] is not None


def test_valid_jpeg_upload_succeeds(client, login_as):
    login_as("root", "password123", "super_admin")
    res = client.post("/api/admin/company-settings/logo",
                       data={"logo": (_jpeg_bytes(), "logo.jpg")}, content_type="multipart/form-data")
    assert res.status_code == 200


def test_valid_webp_upload_succeeds(client, login_as):
    login_as("root", "password123", "super_admin")
    res = client.post("/api/admin/company-settings/logo",
                       data={"logo": (_webp_bytes(), "logo.webp")}, content_type="multipart/form-data")
    assert res.status_code == 200


def test_non_image_file_rejected(client, login_as):
    login_as("root", "password123", "super_admin")
    res = client.post("/api/admin/company-settings/logo",
                       data={"logo": (io.BytesIO(b"not an image, just text"), "fake.png")},
                       content_type="multipart/form-data")
    assert res.status_code == 400


def test_oversized_logo_rejected(client, login_as):
    login_as("root", "password123", "super_admin")
    oversized = io.BytesIO(b"0" * (3 * 1024 * 1024))
    res = client.post("/api/admin/company-settings/logo",
                       data={"logo": (oversized, "big.png")}, content_type="multipart/form-data")
    assert res.status_code == 400
    assert "2MB" in res.get_json()["error"]


def test_no_file_rejected(client, login_as):
    login_as("root", "password123", "super_admin")
    res = client.post("/api/admin/company-settings/logo", data={}, content_type="multipart/form-data")
    assert res.status_code == 400


def test_uploaded_filename_is_server_generated(client, login_as, app):
    login_as("root", "password123", "super_admin")
    res = client.post("/api/admin/company-settings/logo",
                       data={"logo": (_png_bytes(), "../../../etc/passwd.png")},
                       content_type="multipart/form-data")
    assert res.status_code == 200
    with app.app_context():
        from webapp.models.company_settings import CompanySettings
        from webapp.extensions import db
        settings = db.session.get(CompanySettings, 1)
        assert settings.logo_path is not None
        assert ".." not in settings.logo_path
        assert "/" not in settings.logo_path and "\\" not in settings.logo_path
        assert "passwd" not in settings.logo_path
        # server-generated: 32 hex chars (uuid4().hex) + extension
        name_part = settings.logo_path.rsplit(".", 1)[0]
        assert len(name_part) == 32
        int(name_part, 16)  # raises ValueError if not pure hex


def test_replacing_logo_removes_old_file(client, login_as, app):
    login_as("root", "password123", "super_admin")
    client.post("/api/admin/company-settings/logo",
                data={"logo": (_png_bytes(), "first.png")}, content_type="multipart/form-data")
    with app.app_context():
        from webapp.models.company_settings import CompanySettings
        from webapp.extensions import db
        from webapp.services import branding_service
        first_path = branding_service.logo_file_path(db.session.get(CompanySettings, 1))
        assert os.path.exists(first_path)

    client.post("/api/admin/company-settings/logo",
                data={"logo": (_png_bytes(color="blue"), "second.png")}, content_type="multipart/form-data")
    with app.app_context():
        assert not os.path.exists(first_path)


def test_removing_logo_deletes_file_and_clears_path(client, login_as, app):
    login_as("root", "password123", "super_admin")
    client.post("/api/admin/company-settings/logo",
                data={"logo": (_png_bytes(), "logo.png")}, content_type="multipart/form-data")
    with app.app_context():
        from webapp.models.company_settings import CompanySettings
        from webapp.extensions import db
        from webapp.services import branding_service
        logo_path = branding_service.logo_file_path(db.session.get(CompanySettings, 1))
        assert os.path.exists(logo_path)

    res = client.delete("/api/admin/company-settings/logo")
    assert res.status_code == 200
    assert res.get_json()["logo_url"] is None
    with app.app_context():
        assert not os.path.exists(logo_path)


def test_removing_logo_when_none_set_is_rejected(client, login_as):
    login_as("root", "password123", "super_admin")
    res = client.delete("/api/admin/company-settings/logo")
    assert res.status_code == 400


def test_logo_served_via_public_endpoint(client, login_as):
    login_as("root", "password123", "super_admin")
    client.post("/api/admin/company-settings/logo",
                data={"logo": (_png_bytes(), "logo.png")}, content_type="multipart/form-data")
    logo_url = client.get("/api/branding").get_json()["logo_url"]
    res = client.get(logo_url)
    assert res.status_code == 200
    assert res.content_type == "image/png"


def test_no_logo_returns_404(client):
    res = client.get("/api/branding/logo")
    assert res.status_code == 404


# ---------- audit logging ----------

def test_settings_update_is_audited_with_before_after(client, login_as, app):
    login_as("root", "password123", "super_admin")
    client.patch("/api/admin/company-settings", json={"display_name": "Audited Co"})
    with app.app_context():
        from webapp.models.audit_log import AuditLog
        entry = AuditLog.query.filter_by(action="update", entity_type="company_settings").first()
        assert entry is not None
        d = entry.to_dict()
        assert d["before"]["display_name"] == "Daily Figures"
        assert d["after"]["display_name"] == "Audited Co"


def test_logo_upload_is_audited_as_upload(client, login_as, app):
    login_as("root", "password123", "super_admin")
    client.post("/api/admin/company-settings/logo",
                data={"logo": (_png_bytes(), "logo.png")}, content_type="multipart/form-data")
    with app.app_context():
        from webapp.models.audit_log import AuditLog
        entry = AuditLog.query.filter_by(action="logo_upload", entity_type="company_settings").first()
        assert entry is not None


def test_logo_replace_is_audited_as_replace(client, login_as, app):
    login_as("root", "password123", "super_admin")
    client.post("/api/admin/company-settings/logo",
                data={"logo": (_png_bytes(), "first.png")}, content_type="multipart/form-data")
    client.post("/api/admin/company-settings/logo",
                data={"logo": (_png_bytes(color="blue"), "second.png")}, content_type="multipart/form-data")
    with app.app_context():
        from webapp.models.audit_log import AuditLog
        entry = AuditLog.query.filter_by(action="logo_replace", entity_type="company_settings").first()
        assert entry is not None


def test_logo_remove_is_audited(client, login_as, app):
    login_as("root", "password123", "super_admin")
    client.post("/api/admin/company-settings/logo",
                data={"logo": (_png_bytes(), "logo.png")}, content_type="multipart/form-data")
    client.delete("/api/admin/company-settings/logo")
    with app.app_context():
        from webapp.models.audit_log import AuditLog
        entry = AuditLog.query.filter_by(action="logo_remove", entity_type="company_settings").first()
        assert entry is not None


# ---------- branding display ----------

def test_public_branding_reflects_configured_name_and_logo(client, login_as):
    login_as("root", "password123", "super_admin")
    client.patch("/api/admin/company-settings", json={"display_name": "Shown Co"})
    client.post("/api/admin/company-settings/logo",
                data={"logo": (_png_bytes(), "logo.png")}, content_type="multipart/form-data")
    data = client.get("/api/branding").get_json()
    assert data["display_name"] == "Shown Co"
    assert data["logo_url"] is not None
    # never exposes private fields
    assert "email" not in data and "address" not in data and "phone" not in data


# ---------- export integration ----------

def test_csv_export_has_no_branding_rows():
    from webapp.services.export_service import build_csv
    out = build_csv(title="My Report", filters={}, generated_by="alice",
                     columns=[("a", "A")], rows=[{"a": "1"}])
    lines = out.strip().splitlines()
    assert lines[0] == "My Report"  # unchanged: title is always the first line, never a company-name row


def test_xlsx_export_includes_company_header_footer_and_logo(app, login_as, client):
    login_as("root", "password123", "super_admin")
    client.patch("/api/admin/company-settings", json={
        "display_name": "Branded Co", "report_footer_text": "Thanks for your business.",
    })
    client.post("/api/admin/company-settings/logo",
                data={"logo": (_png_bytes(), "logo.png")}, content_type="multipart/form-data")

    with app.app_context():
        from webapp.services import branding_service
        from webapp.services.export_service import build_xlsx
        import openpyxl

        out = build_xlsx(title="My Report", filters={}, generated_by="alice",
                          columns=[("a", "A")], rows=[{"a": "1"}], **branding_service.export_kwargs())
        wb = openpyxl.load_workbook(io.BytesIO(out))
        ws = wb.active
        assert ws.oddHeader.left.text == "Branded Co"
        assert ws.oddFooter.left.text == "Thanks for your business."
        assert len(ws._images) == 1
        # row 1 is still the title, unaffected by branding — no inserted row
        assert ws["A1"].value == "My Report"


def test_xlsx_export_without_branding_configured_has_no_logo_image(app):
    with app.app_context():
        from webapp.services import branding_service
        from webapp.services.export_service import build_xlsx
        import openpyxl

        out = build_xlsx(title="My Report", filters={}, generated_by="alice",
                          columns=[("a", "A")], rows=[{"a": "1"}], **branding_service.export_kwargs())
        wb = openpyxl.load_workbook(io.BytesIO(out))
        ws = wb.active
        assert ws.oddHeader.left.text == "Daily Figures"
        assert len(ws._images) == 0
        assert ws["A1"].value == "My Report"


def test_pdf_export_with_branding_still_produces_valid_pdf(app, login_as, client):
    login_as("root", "password123", "super_admin")
    client.patch("/api/admin/company-settings", json={
        "display_name": "Branded Co", "report_footer_text": "Thanks for your business.",
    })
    client.post("/api/admin/company-settings/logo",
                data={"logo": (_png_bytes(), "logo.png")}, content_type="multipart/form-data")

    with app.app_context():
        from webapp.services import branding_service
        from webapp.services.export_service import build_pdf
        out = build_pdf(title="My Report", filters={}, generated_by="alice",
                         columns=[("a", "A")], rows=[{"a": "1"}], **branding_service.export_kwargs())
        assert out.startswith(b"%PDF")
        assert len(out) > 100


def test_export_route_still_works_with_branding_configured(client, login_as):
    login_as("root", "password123", "super_admin")
    client.patch("/api/admin/company-settings", json={"display_name": "Branded Co"})
    client.post("/api/admin/company-settings/logo",
                data={"logo": (_png_bytes(), "logo.png")}, content_type="multipart/form-data")

    product = client.post("/api/admin/products", json={"name": "Branding Export Product"}).get_json()
    client.post(f"/api/admin/products/{product['id']}/packaging-rules", json={
        "cartons_to_packs": 10, "packs_to_pieces": 10,
    })
    for fmt, mimetype in [
        ("csv", "text/csv"),
        ("xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        ("pdf", "application/pdf"),
    ]:
        res = client.get(f"/api/daily-figures/export.{fmt}")
        assert res.status_code == 200
        assert res.mimetype == mimetype


# ---------- migration ----------

def test_company_settings_migration_up_and_down(tmp_path, monkeypatch):
    import sqlite3

    db_path = tmp_path / "company_settings_migration_test.db"
    monkeypatch.setenv("DB_PATH", str(db_path))
    monkeypatch.setenv("SECRET_KEY", "test-secret")
    monkeypatch.delenv("SUPERADMIN_USERNAME", raising=False)
    monkeypatch.delenv("SUPERADMIN_PASSWORD", raising=False)

    from webapp import create_app
    from flask_migrate import downgrade, upgrade

    flask_app = create_app()
    with flask_app.app_context():
        upgrade()

    conn = sqlite3.connect(db_path)
    row = conn.execute("SELECT display_name, logo_path FROM company_settings WHERE id = 1").fetchone()
    conn.close()
    assert row == ("Daily Figures", None)

    with flask_app.app_context():
        downgrade(revision="2a904d1ebe3b")  # this migration's down_revision

    conn = sqlite3.connect(db_path)
    remaining = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}
    conn.close()
    assert "company_settings" not in remaining
    assert "entries" in remaining
