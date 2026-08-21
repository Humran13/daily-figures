"""
Shared report formatting for every export in the app (dispatch
transactions, daily figures, product movement, customer history,
date-range summaries, filtered search results). One place implements
"title + applied filters + generated date/time + user + headers + totals",
so every export looks and behaves the same way, and quantities are always
written out as exact cartons/packs/pieces columns — never flattened into a
rounded decimal.

`company_settings` is an optional CompanySettings instance (see
webapp/models/company_settings.py) — routes fetch it via
branding_service.get_settings() and pass it through to build_export().
It's kept as an explicit parameter here, never fetched by this module
itself, so these builder functions stay pure and testable without a Flask
app/DB context (see tests/test_export_service.py, which calls them
directly). CSV never receives it — it must stay pure, machine-readable
data with no decorative branding rows.
"""
import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font
from fpdf import FPDF

from webapp.services.business_calendar import format_kampala_report_timestamp


def _generated_at_str():
    """
    User-facing "Generated ..." header for every export format — Africa/
    Kampala local time, labeled EAT, via the one centralized Kampala
    helper (business_calendar.py). Previously rendered raw server UTC
    (e.g. "2026-08-21 06:17 UTC") directly into every report regardless
    of format — the exact bug this fixes; every export's own on-disk
    canonical storage is untouched, this only changes what's PRINTED on
    the generated-at line.
    """
    return format_kampala_report_timestamp()


def _filters_line(filters):
    active = {k: v for k, v in (filters or {}).items() if v not in (None, "", [])}
    if not active:
        return "Filters: none"
    return "Filters: " + "; ".join(f"{k}={v}" for k, v in active.items())


def build_csv(*, title, filters, generated_by, columns, rows, totals=None):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([title])
    writer.writerow([f"Generated {_generated_at_str()} by {generated_by}"])
    writer.writerow([_filters_line(filters)])
    writer.writerow([])
    writer.writerow([label for _, label in columns])
    for row in rows:
        writer.writerow([row.get(key, "") for key, _ in columns])
    if totals:
        writer.writerow([totals.get(key, "") for key, _ in columns])
    return buf.getvalue()


def build_xlsx(*, title, filters, generated_by, columns, rows, totals=None, company_settings=None, logo_path=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Print header/footer metadata, not data rows — company branding never
    # shifts any existing row index (row 1 is always the report title).
    if company_settings:
        ws.oddHeader.left.text = company_settings.display_name
        if company_settings.report_footer_text:
            ws.oddFooter.left.text = company_settings.report_footer_text

    bold = Font(bold=True)
    ws.append([title]); ws["A1"].font = bold
    ws.append([f"Generated {_generated_at_str()} by {generated_by}"])
    ws.append([_filters_line(filters)])
    ws.append([])

    header_row_idx = ws.max_row + 1
    ws.append([label for _, label in columns])
    for cell in ws[header_row_idx]:
        cell.font = bold

    for row in rows:
        ws.append([row.get(key, "") for key, _ in columns])

    if totals:
        total_row_idx = ws.max_row + 1
        ws.append([totals.get(key, "") for key, _ in columns])
        for cell in ws[total_row_idx]:
            cell.font = bold

    for i, (_, label) in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(row=header_row_idx, column=i).column_letter].width = max(12, len(label) + 2)

    if logo_path:
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(logo_path)
            img.width, img.height = 60, 60  # keep a large source logo from dwarfing the sheet
            anchor_col = chr(ord('A') + len(columns) + 1)  # clear of the data columns
            ws.add_image(img, f"{anchor_col}1")
        except Exception:
            pass  # never fail an export over a corrupt/unreadable logo file

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf(*, title, filters, generated_by, columns, rows, totals=None, company_settings=None, logo_path=None):
    pdf = FPDF(orientation="L" if len(columns) > 6 else "P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    if logo_path:
        try:
            pdf.image(logo_path, x=10, y=8, w=20)
            pdf.set_xy(10, 30)
        except Exception:
            pass  # never fail an export over a corrupt/unreadable logo file

    if company_settings:
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 6, company_settings.display_name, new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 8, title, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, f"Generated {_generated_at_str()} by {generated_by}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, _filters_line(filters), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    table_rows = [[label for _, label in columns]]
    table_rows += [[str(row.get(key, "")) for key, _ in columns] for row in rows]
    if totals:
        table_rows.append([str(totals.get(key, "")) for key, _ in columns])

    pdf.set_font("Helvetica", "", 8)
    with pdf.table(text_align="LEFT") as table:
        for i, data_row in enumerate(table_rows):
            row = table.row()
            for datum in data_row:
                row.cell(datum)

    if company_settings and company_settings.report_footer_text:
        pdf.ln(4)
        pdf.set_font("Helvetica", "I", 8)
        pdf.multi_cell(0, 4, company_settings.report_footer_text)

    return bytes(pdf.output())


FORMAT_BUILDERS = {"csv": build_csv, "xlsx": build_xlsx, "pdf": build_pdf}
MIME_TYPES = {
    "csv": "text/csv",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}


def build_export(fmt, **kwargs):
    if fmt not in FORMAT_BUILDERS:
        raise ValueError(f"unsupported export format '{fmt}' — must be one of {list(FORMAT_BUILDERS)}")
    if fmt == "csv":
        # CSV never takes branding kwargs — strip them if a caller passed
        # company_settings/logo_path uniformly to build_export() for every
        # format.
        kwargs = {k: v for k, v in kwargs.items() if k not in ("company_settings", "logo_path")}
    return FORMAT_BUILDERS[fmt](**kwargs)
